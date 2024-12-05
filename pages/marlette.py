import base64
from io import StringIO
import io
from pathlib import Path
from typing import List, Tuple
import streamlit as st
import os
from openpyxl import load_workbook
import zipfile
from fill_from_excel import fill_forms, get_form_fillers
from pydantic import BaseModel
from datetime import datetime

st.set_page_config(page_icon="⭐")

class FileContent(BaseModel):
    name: str
    content: bytes


TMP_DIR_NAME = "tmp"
TPL_DIR_PATH = "assets/templates/blank"
tpl_files: List[FileContent] = []
ds_file: FileContent | None = None


def create_tmp_dir(tmp_dir_name: str):
    try:
        os.mkdir(tmp_dir_name)
        print(f"Directory '{tmp_dir_name}' created successfully.")
    except FileExistsError:
        print(f"Directory '{tmp_dir_name}' already exists.")


@st.fragment
def select_templates(tpl_files: List[FileContent]):
    tpl_files.clear()
    b = st.toggle("Je veux utiliser des modèles de documents personnalisés", value=False)
    if b:
        uploaded_files = st.file_uploader(
            "Déposez les modèles de documents personnalisés au format `pdf` ou `docx` ci-dessous :",
            accept_multiple_files=True,
            type=["pdf", "docx"],
        )
        if uploaded_files is not None and len(uploaded_files) > 0:
            for uploaded_file in uploaded_files:
                tpl_files.append(
                    FileContent(name=uploaded_file.name, content=uploaded_file.read())
                )
    else:
        tpl_dir = Path(TPL_DIR_PATH)
        for tpl_filename in os.listdir(TPL_DIR_PATH):
            tpl_filepath = tpl_dir.joinpath(tpl_filename)
            file_content = open(str(tpl_filepath), "rb").read()
            tpl_files.append(FileContent(name=tpl_filename, content=file_content))
        if len(tpl_files) > 0:
            with st.expander(
                "Liste des modèles de documents sélectionnés par défaut", expanded=True
            ):
                tpl_dir = Path(TPL_DIR_PATH)
                md_list = ""
                for tpl_file in tpl_files:
                    md_list += f"1. {tpl_file.name}\n"
                st.markdown(md_list)

    if len(tpl_files) > 0:
        with io.BytesIO() as buffer:
            with zipfile.ZipFile(buffer, "w") as zip:
                for tpl_file in tpl_files:
                    zip.writestr(tpl_file.name, tpl_file.content)
            buffer.seek(0)

            d = st.download_button(
                label="Télécharger",
                data=buffer,
                file_name="templates.zip",
                type="secondary",
                help="Téléchargez l'ensemble des modèles de documents sélectionnés dans une archive ZIP.",
            )
    else:
        st.warning(
            "Veuillez déposer vos modèles de documents personnalisés pour continuer."
        )


@st.fragment
def select_datasheet(ds_file: FileContent | None) -> FileContent | None:
    ds_file = None
    b = st.toggle("Je veux utiliser la fiche de données par défaut", value=False)
    if not b:
        uploaded_file = st.file_uploader(
            "Déposer la fiche de données au format `.xlsx` ci-dessous :", type=["xlsx"]
        )
        if uploaded_file is not None:
            ds_file = FileContent(name=uploaded_file.name, content=uploaded_file.read())
    else:
        file_name = "fiche.xlsx"
        file_content = open(f"assets/{file_name}", "rb").read()
        ds_file = FileContent(name=file_name, content=file_content)

    if ds_file is not None:
        d = st.download_button(
            label="Télécharger",
            data=ds_file.content,
            file_name=ds_file.name,
            type="secondary",
            help="Téléchargez la fiche de données sélectionnée.",
        )
    else:
        st.warning("Veuillez déposer votre fiche de données pour continuer.")

    return ds_file


create_tmp_dir(TMP_DIR_NAME)


st.title("MARlette ⭐")
st.caption("[Pour vos dossiers *MaPrimeRénov' Parcours Accompagné (MPRA)*]")

st.write(
    """
En 1 clic **MARlette** prépare vos formulaires ANAH à partir d'une simple fiche de suivi au format excel 🚀. 
"""
)
st.caption("Crédits : [Lionel](https://www.linkedin.com/in/lionel-dupeloux/)")

st.header("1. Déposer vos fichiers")
st.subheader("A. Sélectionnez vos modèles de documents  `pdf`  `docx`", divider="gray")

st.caption(
    """Utilisez les modèles par défaut ou chargez vos propres modèles de documents, \
    par exemple si vous voulez utiliser des formulaires pré-signés ou le modèle de contrat de votre entreprise."""
)

st.caption(
    """A tout moment vous pouvez téléchargez les modèles sélectionnés pour comprendre comment fonctionne **MARlette ⭐** et vérifier votre travail."""
)


select_templates(tpl_files)

st.subheader("B. Sélectionnez votre fiche de données `xlsx`", divider="gray")
st.caption(
    """Uploadez votre fiche de données ou utilisez le modèle par défaut pour comprendre comment fonctionne **MARlette ⭐**."""
)

ds_file = select_datasheet(ds_file)


st.header("2. Générer vos documents")
st.caption(
    """Quand vous êtes prêts, cliquez sur le bouton `Générer` : en quelques secondes **MARlette ⭐** rempli vos modèles de documents et vous les propose au téléchargement 🤩."""
)

b1 = st.button(
    "Générer",
    type="primary",
    use_container_width=True,
    help="Cliquez pour remplir les modèles de documents avec les données de la fiche.",
)
#
if b1:
    if tpl_files is None or len(tpl_files) == 0:
        st.toast(
            "❌ Aucun modèle de document n'a été sélectionné. Uploadez vos propores modèles ou utilisez ceux proposés par défaut."
        )
    elif ds_file is None:
        st.toast(
            "❌ Aucune fiche de donnée n'a été sélectionnée. Uploadez votre propre fiche de données ou utilisez celle proposée par défaut."
        )
    else:
        # wb = load_workbook(ds_file[1])
        # st.write(wb.sheetnames)
        with st.spinner("La génération des documents en cours ..."):
            form_fillers = get_form_fillers(io.BytesIO(ds_file.content))
            output_filepaths = fill_forms(
                form_fillers,
                template_base_dir="assets/templates/cccps",
                output_base_dir=TMP_DIR_NAME,
            )

            print(len(output_filepaths))
            tmp_files = os.listdir(TMP_DIR_NAME)
            zip_filenames = []
            for p in output_filepaths:
                if p.name in tmp_files:
                    zip_filenames.append(p.name)
                    print(p.name)

            with io.BytesIO() as buffer:
                with zipfile.ZipFile(buffer, "w") as zip:
                    p = Path(ds_file.name)
                    ts = datetime.now().strftime("%Y%m%dT%H%M%S")
                    zip.writestr(f"{p.stem}_{ts}{p.suffix}", ds_file.content)
                    for filename in zip_filenames:
                        file_path = Path(TMP_DIR_NAME).joinpath(filename)
                        file_arcname = f"{filename}"
                        zip.write(file_path, arcname=file_arcname)
                buffer.seek(0)

                b2 = st.download_button(
                    label="Télécharger",
                    data=buffer,
                    file_name="forms.zip",
                    type="primary",
                    help="Cliquez pour télécharger les documents finaux, remplis avec les données de la fiche.",
                    use_container_width=True,
                )
            st.toast(
                "Vos documents sont remplis et prêt à être téléchargés !", icon="🚀"
            )
